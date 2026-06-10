import os
import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from crewai.tools import tool

# Ensure tasks.xlsx is created in the same directory as this script
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "tasks.xlsx")
SHEET_NAME = "TaskQueue"

def init_excel():
    """Initializes the Excel file with correct headers if it doesn't exist."""
    headers = [
        "Task_ID", "Task_Type", "Description", "Course", 
        "Priority", "Assigned_Agent", "Status", "Created_At", "Notes"
    ]
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(headers)
        wb.save(EXCEL_FILE)
        print(f"Created {EXCEL_FILE} with sheet {SHEET_NAME} and headers.")
    else:
        # Check if the sheet exists, if not, create it
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(title=SHEET_NAME)
            ws.append(headers)
            wb.save(EXCEL_FILE)
            print(f"Added sheet {SHEET_NAME} to existing {EXCEL_FILE}.")

def get_next_task_id() -> str:
    """Reads the Excel file and returns the next Task_ID (e.g., 'T001', 'T002')."""
    init_excel()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
        if df.empty or "Task_ID" not in df.columns or df["Task_ID"].isnull().all():
            return "T001"
        
        # Extract numeric parts and find max
        task_ids = df["Task_ID"].dropna().astype(str).tolist()
        numeric_ids = []
        for tid in task_ids:
            if tid.startswith("T"):
                try:
                    # Strip out any space and convert to int
                    clean_id = tid[1:].strip()
                    numeric_ids.append(int(clean_id))
                except ValueError:
                    pass
        if not numeric_ids:
            return "T001"
        next_num = max(numeric_ids) + 1
        return f"T{next_num:03d}"
    except Exception as e:
        print(f"Error reading next task ID: {e}. Defaulting to T001.")
        return "T001"

def write_task_to_excel_func(
    task_type: str,
    description: str,
    course: str,
    priority: str = "High",
    assigned_agent: str = "TA",
    status: str = "Pending",
    notes: str = ""
) -> str:
    """
    Core function to write a task record to tasks.xlsx.
    """
    try:
        init_excel()
        task_id = get_next_task_id()
        created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Load the workbook using openpyxl for reliable appending
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]
        
        new_row = [
            task_id,
            task_type,
            description,
            course,
            priority,
            assigned_agent,
            status,
            created_at,
            notes
        ]
        ws.append(new_row)
        wb.save(EXCEL_FILE)
        return f"SUCCESS: Task {task_id} successfully created and written to Excel."
    except Exception as e:
        return f"ERROR: Failed to write task to Excel: {str(e)}"

@tool("Excel Writer Tool")
def write_task_to_excel(
    task_type: str,
    description: str,
    course: str,
    priority: str = "High",
    assigned_agent: str = "TA",
    status: str = "Pending",
    notes: str = ""
) -> str:
    """
    Writes a new task record to the tasks.xlsx file under the TaskQueue sheet.
    Automatically generates a unique Task_ID (e.g., T001) and sets the Created_At timestamp.
    
    Args:
        task_type (str): The classified category of the task (e.g. Quiz, Assignment, Flashcards, Study Guide, Summary, Exam).
        description (str): Full explanation of the request.
        course (str): The course or subject name (e.g., Machine Learning).
        priority (str): Priority level (High, Medium, Low). Defaults to High.
        assigned_agent (str): The agent assigned to handle the task. Defaults to TA.
        status (str): The current status of the task. Defaults to Pending.
        notes (str): Additional options/details extracted (e.g. "20 MCQ questions").
    """
    return write_task_to_excel_func(
        task_type=task_type,
        description=description,
        course=course,
        priority=priority,
        assigned_agent=assigned_agent,
        status=status,
        notes=notes
    )
